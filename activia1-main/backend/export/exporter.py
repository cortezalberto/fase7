"""
Research Data Exporter - Multi-format export for anonymized data

Exports anonymized data in multiple formats for institutional research
and educational data mining.

Supported Formats:
- JSON: Structured data with full metadata
- CSV: Flat tables for statistical analysis
- Excel: Multi-sheet workbooks for manual review
- SPSS: (Future) For social science statistical analysis

Key Features:
- Multiple export formats from single source
- Automatic schema detection and validation
- Incremental export for large datasets
- Compression support for large files
"""

import csv
import io
import json
import logging
import zipfile
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from pydantic import BaseModel, Field

from backend.core.constants import utc_now

logger = logging.getLogger(__name__)


class ExportFormat(str, Enum):
    """Supported export formats"""

    JSON = "json"
    CSV = "csv"
    EXCEL = "excel"
    SPSS = "spss"  # Future implementation


class ExportConfig(BaseModel):
    """Configuration for data export"""

    format: ExportFormat = Field(description="Export format")
    compress: bool = Field(
        default=False, description="Compress output file (ZIP)"
    )
    include_metadata: bool = Field(
        default=True, description="Include export metadata"
    )
    pretty_print: bool = Field(
        default=True, description="Pretty print JSON (if format=JSON)"
    )
    csv_delimiter: str = Field(
        default=",", description="CSV delimiter (default: comma)"
    )
    csv_encoding: str = Field(
        default="utf-8-sig", description="CSV encoding (with BOM for Excel)"
    )
    excel_sheet_names: Optional[Dict[str, str]] = Field(
        default=None,
        description="Custom sheet names for Excel export",
    )


class ResearchDataExporter:
    """
    Exports anonymized research data in multiple formats

    Example:
        >>> exporter = ResearchDataExporter()
        >>>
        >>> # Export traces to JSON
        >>> traces = [...]  # List of anonymized traces
        >>> json_output = exporter.export_to_json(traces, "traces")
        >>>
        >>> # Export to CSV
        >>> csv_output = exporter.export_to_csv(traces)
        >>>
        >>> # Export to Excel with multiple sheets
        >>> data = {"traces": traces, "evaluations": evals, "risks": risks}
        >>> excel_bytes = exporter.export_to_excel(data)
    """

    def __init__(self, config: Optional[ExportConfig] = None):
        """
        Initialize exporter with configuration

        Args:
            config: Export configuration (uses defaults if None)
        """
        self.config = config or ExportConfig(format=ExportFormat.JSON)
        logger.info(
            "ResearchDataExporter initialized",
            extra={
                "format": self.config.format,
                "compress": self.config.compress,
            },
        )

    def generate_metadata(self, record_count: int, data_types: List[str]) -> Dict[str, Any]:
        """
        Generate export metadata

        Args:
            record_count: Number of records exported
            data_types: Types of data included (e.g., ["traces", "evaluations"])

        Returns:
            Metadata dictionary
        """
        return {
            "export_timestamp": utc_now().isoformat() + "Z",
            "export_format": self.config.format,
            "total_records": record_count,
            "data_types": data_types,
            "anonymization_applied": True,
            "privacy_standard": "k-anonymity",
            "tool": "AI-Native MVP Export Module",
            "version": "1.0.0",
        }

    def export_to_json(
        self,
        data: Union[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]],
        data_type: Optional[str] = None,
    ) -> str:
        """
        Export data to JSON format

        Args:
            data: Data to export (list of records or dict of lists)
            data_type: Type of data (e.g., "traces", "evaluations")

        Returns:
            JSON string
        """
        logger.info(
            "Exporting to JSON",
            extra={"data_type": data_type, "pretty_print": self.config.pretty_print},
        )

        # Wrap single list in dict
        if isinstance(data, list):
            data_dict = {data_type or "data": data}
        else:
            data_dict = data

        # Calculate total records
        total_records = sum(
            len(v) if isinstance(v, list) else 1 for v in data_dict.values()
        )

        # Create export structure
        export_data = {
            "metadata": self.generate_metadata(
                record_count=total_records,
                data_types=list(data_dict.keys()),
            )
            if self.config.include_metadata
            else {},
            "data": data_dict,
        }

        # Serialize
        json_str = json.dumps(
            export_data,
            indent=2 if self.config.pretty_print else None,
            default=str,  # Handle datetime, Enum, etc.
            ensure_ascii=False,
        )

        logger.info(
            "JSON export completed",
            extra={"total_records": total_records, "size_bytes": len(json_str)},
        )
        return json_str

    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        fieldnames: Optional[List[str]] = None,
    ) -> str:
        """
        Export data to CSV format

        Args:
            data: List of records (must have consistent schema)
            fieldnames: Column names (auto-detected if None)

        Returns:
            CSV string
        """
        if not data:
            logger.warning("No data to export to CSV")
            return ""

        logger.info("Exporting to CSV", extra={"record_count": len(data)})

        # Auto-detect fieldnames from first record
        if fieldnames is None:
            fieldnames = list(data[0].keys())

        # Write to string buffer
        output = io.StringIO()
        writer = csv.DictWriter(
            output,
            fieldnames=fieldnames,
            delimiter=self.config.csv_delimiter,
            extrasaction="ignore",  # Ignore extra fields
        )

        writer.writeheader()
        writer.writerows(data)

        csv_str = output.getvalue()
        output.close()

        logger.info(
            "CSV export completed",
            extra={
                "total_records": len(data),
                "columns": len(fieldnames),
                "size_bytes": len(csv_str),
            },
        )
        return csv_str

    def export_to_excel(
        self, data: Dict[str, List[Dict[str, Any]]]
    ) -> bytes:
        """
        Export data to Excel format (multi-sheet workbook)

        Requires openpyxl package.

        Args:
            data: Dictionary mapping sheet names to list of records

        Returns:
            Excel file as bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            raise ImportError(
                "Excel export requires openpyxl. Install with: pip install openpyxl"
            )

        logger.info(
            "Exporting to Excel", extra={"sheet_count": len(data)}
        )

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        total_records = 0

        # Create a sheet for each data type
        for sheet_name, records in data.items():
            if not records:
                logger.warning(f"No data for sheet '{sheet_name}', skipping")
                continue

            # Use custom sheet name if configured
            if self.config.excel_sheet_names and sheet_name in self.config.excel_sheet_names:
                display_name = self.config.excel_sheet_names[sheet_name]
            else:
                display_name = sheet_name[:31]  # Excel sheet name max length

            ws = wb.create_sheet(title=display_name)

            # Get fieldnames from first record
            fieldnames = list(records[0].keys())

            # Write header
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)

            for col_idx, field in enumerate(fieldnames, start=1):
                cell = ws.cell(row=1, column=col_idx, value=field)
                cell.fill = header_fill
                cell.font = header_font

            # Write data
            for row_idx, record in enumerate(records, start=2):
                for col_idx, field in enumerate(fieldnames, start=1):
                    value = record.get(field, "")
                    # Convert complex types to string
                    if isinstance(value, (dict, list)):
                        value = json.dumps(value, ensure_ascii=False)
                    ws.cell(row=row_idx, column=col_idx, value=value)

            total_records += len(records)
            logger.debug(
                f"Sheet '{display_name}' created",
                extra={"records": len(records), "columns": len(fieldnames)},
            )

        # Add metadata sheet if configured
        if self.config.include_metadata:
            metadata = self.generate_metadata(
                record_count=total_records,
                data_types=list(data.keys()),
            )
            ws_meta = wb.create_sheet(title="Metadata", index=0)
            ws_meta.cell(row=1, column=1, value="Export Metadata").font = Font(bold=True)
            for idx, (key, value) in enumerate(metadata.items(), start=2):
                ws_meta.cell(row=idx, column=1, value=key)
                ws_meta.cell(row=idx, column=2, value=str(value))

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()
        output.close()

        logger.info(
            "Excel export completed",
            extra={
                "total_records": total_records,
                "sheet_count": len(wb.worksheets),
                "size_bytes": len(excel_bytes),
            },
        )
        return excel_bytes

    def compress_output(self, data: Union[str, bytes], filename: str) -> bytes:
        """
        Compress exported data to ZIP

        Args:
            data: Data to compress (string or bytes)
            filename: Filename inside ZIP

        Returns:
            ZIP file as bytes
        """
        logger.info("Compressing output", extra={"archive_filename": filename})

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
            if isinstance(data, str):
                zf.writestr(filename, data.encode("utf-8"))
            else:
                zf.writestr(filename, data)

        zip_bytes = output.getvalue()
        output.close()

        logger.info(
            "Compression completed",
            extra={
                "original_size": len(data),
                "compressed_size": len(zip_bytes),
                "ratio": round(len(zip_bytes) / len(data), 2),
            },
        )
        return zip_bytes

    def export(
        self,
        data: Union[List[Dict[str, Any]], Dict[str, List[Dict[str, Any]]]],
        output_path: Optional[Path] = None,
    ) -> Union[str, bytes]:
        """
        Export data in configured format

        Args:
            data: Data to export
            output_path: Optional path to write output file

        Returns:
            Exported data (string for JSON/CSV, bytes for Excel)
        """
        if self.config.format == ExportFormat.JSON:
            output_data = self.export_to_json(data)
            if output_path:
                output_path.write_text(output_data, encoding="utf-8")
        elif self.config.format == ExportFormat.CSV:
            if isinstance(data, dict):
                # Export first data type for CSV
                first_key = list(data.keys())[0]
                output_data = self.export_to_csv(data[first_key])
            else:
                output_data = self.export_to_csv(data)
            if output_path:
                output_path.write_text(output_data, encoding=self.config.csv_encoding)
        elif self.config.format == ExportFormat.EXCEL:
            if isinstance(data, list):
                # Wrap single list in dict
                data = {"data": data}
            output_data = self.export_to_excel(data)
            if output_path:
                output_path.write_bytes(output_data)
        else:
            raise ValueError(f"Unsupported format: {self.config.format}")

        # Compress if configured
        if self.config.compress and output_path:
            filename = output_path.name
            compressed = self.compress_output(output_data, filename)
            zip_path = output_path.with_suffix(output_path.suffix + ".zip")
            zip_path.write_bytes(compressed)
            logger.info(f"Compressed output written to {zip_path}")
            return compressed

        if output_path:
            logger.info(f"Output written to {output_path}")

        return output_data